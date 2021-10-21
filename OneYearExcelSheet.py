import random # library for random number generation
from openpyxl import Workbook, load_workbook # library for loading Excel file
from openpyxl.utils import get_column_letter # library for accessing Excel file

bloodTypePercentage = [26.65, 23.72, 6.06, 43.57] # A, B, AB, O

wb = load_workbook('2011 Taiwan Blood Supply Demand Modeling.xlsx', data_only=True) # loads the desired Excel file specified by file name 
ws = wb['daily'] # navigates to the correct sheet inside Excel


def demand(): # proceses weekly data, in our case, we do not change it and merely copy it over
    for i in range(0, 52): # 52 weeks
        row = (i*7)+2
        col = 'M'
        originalDemand = ws['L2'].value
        change = 1 + (random.uniform(-0, 0)/100) # NOT randomizing here, therefore, 0% change
        ws[col + str(row)] = originalDemand * change # puts in new cell in same row
def supply(): # proceses weekly data, in our case, we do not change it and merely copy it over
    for i in range(0, 52):
        row = (i*7)+2
        col = 'K'
        originalSupply = ws['J' + str(row)].value
        change = 1 + (random.uniform(-0, 0)/100) # NOT randomizing here, therefore, 0% change
        ws[col + str(row)] = originalSupply * change # puts in new cell in same row

def bloodTypeSupply(): # multiplies the weekly supply by standard blood type ratios (from bloodTypePercentage list), then proceeds to fluctuate by +-supplyMaxVariationPercentage
    for i in range(0, 52): # 52 weeks
        row = (i*7)+2
        for j in range(0, 4):
            col = chr(78 + j) # start with upper case N
            typeChange = 0
            originalSupply = ws['K' + str(row)].value
            originalTypeSupply = originalSupply * (bloodTypePercentage[j]/100) # first multiply by standard blood ratios
            typeChange = 1 + (random.uniform(-supplyMaxVariationPercentage, supplyMaxVariationPercentage)/100) # then vary by +- supplyMaxVariationPercentage
            ws[col + str(row)] = originalTypeSupply * typeChange 

def bloodTypeDemand(): # multiplies the weekly supply by standard blood type ratios (from bloodTypePercentage list), then proceeds to fluctuate by +-supplyMaxVariationPercentage
    for i in range(0, 52): # 52 weeks
        row = (i*7)+2
        for j in range(0, 4):
            col = chr(83 + j) # start with upper case S
            typeChange = 0
            originalDemand = ws['M' + str(row)].value
            originalTypeDemand = originalDemand * (bloodTypePercentage[j]/100) # first multiply by standard blood ratios
            typeChange = 1 + (random.uniform(-demandMaxVariationPercentage, demandMaxVariationPercentage)/100) # then vary by +- demandMaxVariationPercentage
            ws[col + str(row)] = originalTypeDemand * typeChange
            

def bloodTypeDifference():# calculates supply - demand for each bloodtype per week, returns # of weeks our project is useful
    sumPerYear = 0 # number of weeks our project is useful
    for i in range(0, 52): # 52 weeks
        row = (i*7)+2
        difference = [] # list of differences for each blood type [A, B, AB, O]
        for j in range(0, 4):
            col = 87 + j # start with upper case X
            col2 = col - 9 # index of supply column
            col3 = col - 4 # index of demand column
            
            result = ws[chr(col2) + str(row)].value - ws[chr(col3) + str(row)].value # supply - demand
            ws[chr(col) + str(row)] = result
            difference.append(result) # A, B, AB, O
        A = difference[0]
        B = difference[1]
        AB = difference[2]
        O = difference[3]
        # conditions where our project will be useful, add 1 to sumPerYear if useful for this week
        if O < 0 and (A>0 or B>0 or AB>0):
            sumPerYear += 1
        elif (O < 0 and A < 0) and (abs(O) < abs(A)) and (B>0 or AB>0):
            sumPerYear += 1
        elif (O < 0 and B < 0) and (abs(O) < abs(B)) and (A>0 or AB>0):
            sumPerYear += 1
        elif (A < 0 and O > 0) and (abs(O) < abs(A)) and (B>0 or AB>0):
            sumPerYear += 1
        elif (B < 0 and O > 0) and (abs(O) < abs(B)) and (A>0 or AB>0):
            sumPerYear += 1
        else:
            sumPerYear += 0

    return sumPerYear



demand()
supply()
bloodTypeSupply()
bloodTypeDemand()
bloodTypeDifference()
wb.save('2015 Taiwan Blood Supply Demand Modeling.xlsx')
print("complete")





# for weekNumber in range(52):
#     change = 1 + (random.uniform(-2.5, 2.5)/100)
#     weekDemand.append(demandNum * change)

# print(weekDemand)



